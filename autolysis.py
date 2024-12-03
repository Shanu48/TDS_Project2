# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "httpx",
#     "pandas",
#     "seaborn",
#     "matplotlib",
# ]
# ///


import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Load the workbook
file_path = 'goodreads.xlsx'  # Update with the actual path if needed
workbook = load_workbook(file_path)

# Check available sheet names
print(workbook.sheetnames)

